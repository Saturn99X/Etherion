"""
Production-ready Excel Generation Tool using openpyxl.

Supports:
- Professional Excel workbooks with multiple sheets
- Formatted tables with styles
- Formulas and calculations
- Charts and graphs
- Data validation
- Conditional formatting
- Financial reports and data exports

Library: openpyxl
Output: Excel files (.xlsx) with proper formatting
"""

import os
from typing import Dict, List, Optional, Any, Union
from io import BytesIO
from datetime import datetime
import logging

from .base_file_generator import BaseFileGenerator

logger = logging.getLogger(__name__)


class ExcelGeneratorTool(BaseFileGenerator):
    """
    Professional Excel generation tool using openpyxl.

    Features:
    - Multi-sheet workbooks
    - Formatted tables with styles
    - Charts and visualizations
    - Formulas and calculations
    - Data validation
    - Production-ready output
    """

    def __init__(
        self,
        tenant_id: str,
        agent_id: str,
        job_id: str,
        user_id: Optional[str] = None,
        project_id: Optional[str] = None,
    ):
        """
        Initialize the Excel generator.

        Args:
            tenant_id: Tenant ID for isolation
            agent_id: Agent creating the Excel file
            job_id: Associated job ID
            user_id: Optional user ID
            project_id: GCP project ID
        """
        super().__init__(tenant_id, agent_id, job_id, user_id, project_id)

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        # Define common styles
        self.header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.title_font = Font(name="Arial", size=14, bold=True)
        self.title_alignment = Alignment(horizontal="left", vertical="center")

        self.border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        logger.info(f"Initialized ExcelGeneratorTool for tenant={tenant_id}")

    async def generate_workbook(
        self,
        sheets: List[Dict[str, Any]],
        filename: str = "workbook.xlsx",
        description: str = "",
        tags: Optional[List[str]] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Generate a multi-sheet Excel workbook.

        Args:
            sheets: List of sheet configurations with name, data, and formatting
            filename: Output filename
            description: Description for metadata
            tags: Optional tags
            metadata: Optional additional metadata

        Returns:
            Dictionary containing asset info
        """
        from openpyxl import Workbook
        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Process each sheet
            for sheet_config in sheets:
                sheet_name = sheet_config.get("name", "Sheet1")
                data = sheet_config.get("data", [])
                sheet_type = sheet_config.get("type", "table")
                formatting = sheet_config.get("formatting", {})

                # Create sheet
                ws = wb.create_sheet(title=sheet_name)

                if sheet_type == "table":
                    self._create_table_sheet(ws, data, formatting)
                elif sheet_type == "report":
                    self._create_report_sheet(ws, data, formatting)
                elif sheet_type == "raw":
                    self._create_raw_sheet(ws, data, formatting)

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            # Save asset
            asset_info = await self.save_asset(
                file_bytes=excel_bytes,
                filename=filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                description=description or f"Excel workbook with {len(sheets)} sheets",
                tags=tags or ["excel", "workbook", "generated"],
                metadata={
                    **(metadata or {}),
                    "sheet_count": len(sheets),
                    "sheet_names": [s.get("name") for s in sheets],
                },
            )

            logger.info(f"Successfully generated Excel workbook: {filename}")

            return asset_info

        except Exception as e:
            logger.error(f"Error generating Excel workbook: {e}", exc_info=True)
            raise

    def _create_table_sheet(
        self, ws, data: List[List[Any]], formatting: Dict[str, Any]
    ):
        """
        Create a formatted table sheet.

        Args:
            ws: Worksheet object
            data: Table data (first row should be headers)
            formatting: Formatting configuration
        """
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        if not data:
            return

        # Write data
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply header formatting
                if row_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                    cell.border = self.border_thin
                else:
                    cell.border = self.border_thin
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col_idx in range(1, len(data[0]) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in data:
                if col_idx <= len(row):
                    cell_value = str(row[col_idx - 1])
                    max_length = max(max_length, len(cell_value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply table style if requested
        if formatting.get("table_style", False):
            table_name = formatting.get("table_name", "DataTable")
            table_range = f"A1:{get_column_letter(len(data[0]))}{len(data)}"
            table = Table(displayName=table_name, ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        # Freeze panes (freeze first row)
        if formatting.get("freeze_panes", True):
            ws.freeze_panes = "A2"

    def _create_report_sheet(
        self, ws, data: Dict[str, Any], formatting: Dict[str, Any]
    ):
        """
        Create a formatted report sheet with title, sections, and data.

        Args:
            ws: Worksheet object
            data: Report data with title, sections, and content
            formatting: Formatting configuration
        """
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        current_row = 1

        # Add title
        title = data.get("title", "Report")
        ws.cell(row=current_row, column=1, value=title).font = self.title_font
        current_row += 2

        # Add metadata
        metadata = data.get("metadata", {})
        for key, value in metadata.items():
            ws.cell(row=current_row, column=1, value=f"{key}:")
            ws.cell(row=current_row, column=2, value=str(value))
            current_row += 1

        current_row += 1

        # Add sections
        sections = data.get("sections", [])
        for section in sections:
            section_title = section.get("title", "")
            section_data = section.get("data", [])

            # Section title
            ws.cell(row=current_row, column=1, value=section_title).font = Font(
                name="Arial", size=12, bold=True
            )
            current_row += 1

            # Section data
            if section_data:
                for row_data in section_data:
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1

            current_row += 1

        # Auto-adjust column widths
        for col_idx in range(1, 10):  # Adjust first 10 columns
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 15

    def _create_raw_sheet(self, ws, data: List[List[Any]], formatting: Dict[str, Any]):
        """
        Create a raw data sheet without formatting.

        Args:
            ws: Worksheet object
            data: Raw data
            formatting: Formatting configuration
        """
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    async def generate_data_table(
        self,
        data: List[Dict[str, Any]],
        headers: Optional[List[str]] = None,
        filename: str = "data_table.xlsx",
        sheet_name: str = "Data",
        description: str = "",
        tags: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Generate a simple data table Excel file.

        Args:
            data: List of dictionaries representing rows
            headers: Optional list of column headers (inferred from data if not provided)
            filename: Output filename
            sheet_name: Sheet name
            description: Description for metadata
            tags: Optional tags

        Returns:
            Dictionary containing asset info
        """
        try:
            # Infer headers from data if not provided
            if headers is None and data:
                headers = list(data[0].keys())

            # Convert data to rows
            rows = [headers] if headers else []
            for item in data:
                if headers:
                    row = [item.get(h, "") for h in headers]
                else:
                    row = list(item.values())
                rows.append(row)

            # Create workbook
            sheets = [
                {
                    "name": sheet_name,
                    "data": rows,
                    "type": "table",
                    "formatting": {"table_style": True, "freeze_panes": True},
                }
            ]

            return await self.generate_workbook(
                sheets=sheets,
                filename=filename,
                description=description or f"Data table with {len(data)} rows",
                tags=tags or ["excel", "data-table"],
                metadata={"row_count": len(data), "column_count": len(headers or [])},
            )

        except Exception as e:
            logger.error(f"Error generating data table: {e}", exc_info=True)
            raise

    async def generate_financial_report(
        self,
        report_title: str,
        period: str,
        summary_data: Dict[str, float],
        detail_data: List[Dict[str, Any]],
        filename: str = "financial_report.xlsx",
        description: str = "",
        tags: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Generate a financial report Excel file.

        Args:
            report_title: Report title
            period: Reporting period (e.g., "Q1 2025")
            summary_data: Summary metrics as key-value pairs
            detail_data: Detailed transaction data
            filename: Output filename
            description: Description for metadata
            tags: Optional tags

        Returns:
            Dictionary containing asset info
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font
        try:
            # Create workbook
            wb = Workbook()

            # Summary sheet
            summary_sheet = wb.active
            summary_sheet.title = "Summary"

            row = 1
            summary_sheet.cell(
                row=row, column=1, value=report_title
            ).font = self.title_font
            row += 1
            summary_sheet.cell(row=row, column=1, value=f"Period: {period}")
            row += 2

            # Summary metrics
            for key, value in summary_data.items():
                summary_sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
                cell = summary_sheet.cell(row=row, column=2, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "$#,##0.00"
                row += 1

            # Detail sheet
            if detail_data:
                detail_sheet = wb.create_sheet(title="Details")
                headers = list(detail_data[0].keys())

                # Write headers
                for col_idx, header in enumerate(headers, start=1):
                    cell = detail_sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment

                # Write data
                for row_idx, item in enumerate(detail_data, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = item.get(header, "")
                        cell = detail_sheet.cell(
                            row=row_idx, column=col_idx, value=value
                        )
                        # Format currency columns
                        if isinstance(value, (int, float)) and (
                            "amount" in header.lower()
                            or "total" in header.lower()
                            or "price" in header.lower()
                        ):
                            cell.number_format = "$#,##0.00"

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            # Save asset
            asset_info = await self.save_asset(
                file_bytes=excel_bytes,
                filename=filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                description=description or f"Financial report: {report_title}",
                tags=tags or ["excel", "financial-report", "report"],
                metadata={
                    "report_title": report_title,
                    "period": period,
                    "summary_metrics": len(summary_data),
                    "detail_rows": len(detail_data),
                },
            )

            logger.info(f"Successfully generated financial report: {filename}")

            return asset_info

        except Exception as e:
            logger.error(f"Error generating financial report: {e}", exc_info=True)
            raise

    async def generate_chart_workbook(
        self,
        chart_data: Dict[str, Any],
        filename: str = "chart_report.xlsx",
        description: str = "",
        tags: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Generate an Excel workbook with charts.

        Args:
            chart_data: Chart configuration with type, data, and labels
            filename: Output filename
            description: Description for metadata
            tags: Optional tags

        Returns:
            Dictionary containing asset info
        """
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Chart Data"

            # Write data
            data = chart_data.get("data", [])
            labels = chart_data.get("labels", [])
            chart_type = chart_data.get("type", "bar")
            chart_title = chart_data.get("title", "Chart")

            # Write labels and data
            ws.cell(row=1, column=1, value="Category")
            ws.cell(row=1, column=2, value="Value")

            for idx, (label, value) in enumerate(zip(labels, data), start=2):
                ws.cell(row=idx, column=1, value=label)
                ws.cell(row=idx, column=2, value=value)

            # Create chart
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()

            chart.title = chart_title
            chart.style = 10

            # Set data
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(labels) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # Add chart to sheet
            ws.add_chart(chart, "D2")

            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            # Save asset
            asset_info = await self.save_asset(
                file_bytes=excel_bytes,
                filename=filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                description=description or f"Chart report: {chart_title}",
                tags=tags or ["excel", "chart", "visualization"],
                metadata={
                    "chart_title": chart_title,
                    "chart_type": chart_type,
                    "data_points": len(data),
                },
            )

            logger.info(f"Successfully generated chart workbook: {filename}")

            return asset_info

        except Exception as e:
            logger.error(f"Error generating chart workbook: {e}", exc_info=True)
            raise
